"""
GCP Billing SQLite importer.

Usage:
  # Import project master
  python importer.py --projects --excel ../docs/上雲專案預算控管表（GCP 總表）_超新版.xlsx --db billing.db

  # Import billing folder
  python importer.py --billing ../docs/2026/2026-04/AI --month 2026-04 --db billing.db
"""

import argparse
import sqlite3
from datetime import datetime, date
from pathlib import Path

import openpyxl

from db import get_connection, init_db

# ---------------------------------------------------------------------------
# project_master
# ---------------------------------------------------------------------------

_PROJECT_SHEET = "雲端專案控管表"

_PROJECT_COL_MAP = {
    "類別": "category",
    "板塊": "sector",
    "部門": "department",
    "科別": "section",
    "專案名稱": "project_name",
    "專案編號(Project ID)": "project_id",
    "APID": "apid",
    "環境": "environment",
    "成本中心": "cost_center",
}


def import_projects(db_path: str, excel_path: str) -> None:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[_PROJECT_SHEET]

    # Scan for the header row (the one containing the project ID column)
    header_row_num = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if "專案編號(Project ID)" in row:
            header_row_num = i
            break
    if header_row_num is None:
        raise ValueError(f"Cannot find header row in sheet '{_PROJECT_SHEET}'")

    header = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]
    col_index = {name: i for i, name in enumerate(header) if name}

    rows = []
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        project_id = row[col_index["專案編號(Project ID)"]]
        if not project_id:
            continue
        rows.append({
            "project_id":   str(project_id).strip(),
            "apid":         _str(row, col_index, "APID"),
            "category":     _str(row, col_index, "類別"),
            "sector":       _str(row, col_index, "板塊"),
            "department":   _str(row, col_index, "部門"),
            "section":      _str(row, col_index, "科別"),
            "project_name": _str(row, col_index, "專案名稱"),
            "environment":  _str(row, col_index, "環境"),
            "cost_center":  _str(row, col_index, "成本中心"),
        })

    with get_connection(db_path) as conn:
        conn.executemany(
            """
            INSERT INTO project_master
                (project_id, apid, category, sector, department, section, project_name, environment, cost_center)
            VALUES
                (:project_id, :apid, :category, :sector, :department, :section, :project_name, :environment, :cost_center)
            ON CONFLICT(project_id) DO UPDATE SET
                apid         = excluded.apid,
                category     = excluded.category,
                sector       = excluded.sector,
                department   = excluded.department,
                section      = excluded.section,
                project_name = excluded.project_name,
                environment  = excluded.environment,
                cost_center  = excluded.cost_center
            """,
            rows,
        )

    print(f"project_master: upserted {len(rows)} rows")


# ---------------------------------------------------------------------------
# billing (line items + summary)
# ---------------------------------------------------------------------------

_DETAIL_HEADER_ROW = 12
_DETAIL_DATA_START = 13

_BILLING_SHEETS = ["國泰世華", "Apigee"]

_DETAIL_COLS = [
    "project_id", "project_name", "service_description", "service_id",
    "sku_description", "sku_id", "usage_start_date", "usage_end_date",
    "usage_amount", "usage_unit",
    "free_tier", "committed_usage_discount", "committed_usage_discount_dollar_base",
    "discount", "promotion", "subscription_benefit", "sustained_usage_discount",
    "google_cost", "partner_cost", "google_discount", "partner_discount", "total_discount",
]

_DISCOUNT_COLS = {
    "free_tier", "committed_usage_discount", "committed_usage_discount_dollar_base",
    "discount", "promotion", "subscription_benefit", "sustained_usage_discount",
    "google_discount", "partner_discount", "total_discount",
}


def import_billing(db_path: str, folder_path: str, billing_year_month: str, force: bool = False) -> None:
    folder = Path(folder_path).resolve()
    billing_group = folder.name
    imported_at = datetime.now().isoformat(timespec="seconds")

    with get_connection(db_path) as conn:
        if force:
            existing = conn.execute(
                "SELECT id FROM billing_file WHERE billing_group=? AND billing_year_month=?",
                (billing_group, billing_year_month),
            ).fetchone()
            if existing:
                old_id = existing[0]
                conn.execute("DELETE FROM billing_summary WHERE billing_file_id=?", (old_id,))
                conn.execute("DELETE FROM billing_line_item WHERE billing_file_id=?", (old_id,))
                conn.execute("DELETE FROM billing_file WHERE id=?", (old_id,))
                print(f"Force-deleted existing: {billing_group} / {billing_year_month} (id={old_id})")

        try:
            conn.execute(
                "INSERT INTO billing_file (billing_group, billing_year_month, imported_at) VALUES (?, ?, ?)",
                (billing_group, billing_year_month, imported_at),
            )
            billing_file_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        except sqlite3.IntegrityError:
            print(f"Already imported: {billing_group} / {billing_year_month} — skipping (use --force to reimport)")
            return

        xlsx_files = sorted(folder.glob("*.xlsx"))
        if not xlsx_files:
            print(f"No .xlsx files found in {folder}")
            return

        all_line_items = []
        summary_data = None  # [gcp, partner, rate, twd]

        for xlsx_path in xlsx_files:
            if xlsx_path.name.startswith("~$"):
                continue
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)

            has_billing_sheet = any(s in wb.sheetnames for s in _BILLING_SHEETS)
            if not has_billing_sheet:
                continue

            for sheet_name in _BILLING_SHEETS:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                all_line_items.extend(_read_line_items(ws, billing_file_id))

            ws0 = wb[wb.sheetnames[0]]
            candidate = _read_summary(ws0)
            if any(v is not None for v in candidate):
                if summary_data is None:
                    summary_data = list(candidate)
                else:
                    # Accumulate GCP, partner, TWD; keep exchange rate from first file
                    summary_data = [
                        (summary_data[0] or 0) + (candidate[0] or 0),
                        (summary_data[1] or 0) + (candidate[1] or 0),
                        summary_data[2] or candidate[2],
                        (summary_data[3] or 0) + (candidate[3] or 0),
                    ]

        _insert_line_items(conn, all_line_items)

        if summary_data:
            conn.execute(
                """
                INSERT INTO billing_summary
                    (billing_file_id, gcp_usage_total_usd, partner_usage_total_usd, exchange_rate, total_payable_twd)
                VALUES (?, ?, ?, ?, ?)
                """,
                (billing_file_id, *summary_data),
            )

    print(f"billing_file id={billing_file_id}: {billing_group} / {billing_year_month}")
    print(f"  line_items inserted: {len(all_line_items)}")
    print(f"  summary: {summary_data}")


def _read_line_items(ws, billing_file_id: int) -> list:
    rows = []
    for row in ws.iter_rows(min_row=_DETAIL_DATA_START, values_only=True):
        if all(v is None for v in row):
            continue
        item = {"billing_file_id": billing_file_id}
        for i, col in enumerate(_DETAIL_COLS):
            val = row[i] if i < len(row) else None
            if col in ("usage_start_date", "usage_end_date"):
                val = _to_date_str(val)
            elif col in _DISCOUNT_COLS:
                val = float(val) if val is not None else 0.0
            item[col] = val
        rows.append(item)
    return rows


def _insert_line_items(conn: sqlite3.Connection, rows: list) -> None:
    conn.executemany(
        f"""
        INSERT INTO billing_line_item
            (billing_file_id, {', '.join(_DETAIL_COLS)})
        VALUES
            (:billing_file_id, {', '.join(':' + c for c in _DETAIL_COLS)})
        """,
        rows,
    )


def _read_summary(ws) -> tuple:
    def cell(row, col):
        return ws.cell(row=row, column=col).value

    gcp_total   = cell(3, 8)   # H3
    partner     = cell(4, 8)   # H4
    rate        = cell(5, 7)   # G5
    payable_twd = cell(10, 8)  # H10
    return (gcp_total, partner, rate, payable_twd)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _str(row, col_index, key) -> str | None:
    idx = col_index.get(key)
    if idx is None:
        return None
    val = row[idx]
    return str(val).strip() if val is not None else None


def _to_date_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="GCP Billing Importer")
    parser.add_argument("--db", default="billing.db", help="SQLite DB path")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--projects", action="store_true", help="Import project_master")
    group.add_argument("--billing", metavar="FOLDER", help="Billing folder path")

    parser.add_argument("--excel", default="../docs/上雲專案預算控管表（GCP 總表）_超新版.xlsx",
                        help="Project master Excel path (used with --projects)")
    parser.add_argument("--month", metavar="YYYY-MM",
                        help="Billing year-month (used with --billing)")
    parser.add_argument("--force", action="store_true",
                        help="Delete existing record and reimport (used with --billing)")

    args = parser.parse_args()
    init_db(args.db)

    if args.projects:
        import_projects(args.db, args.excel)
    else:
        if not args.month:
            parser.error("--month YYYY-MM is required with --billing")
        import_billing(args.db, args.billing, args.month, force=args.force)


if __name__ == "__main__":
    main()
